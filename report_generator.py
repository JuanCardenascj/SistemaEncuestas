import os
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch

class ReportGenerator:
    
    @staticmethod
    def generar_excel(estadisticas: dict, nombre_archivo: str = "reporte.xlsx"):
        """Genera reporte en formato Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Encuesta"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Encabezado
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Reporte de Encuesta: {estadisticas['titulo_encuesta']}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_align
        
        ws['A2'] = f"Total de respuestas: {estadisticas['total_respuestas']}"
        ws['A2'].font = Font(bold=True)
        
        # Espacio
        ws.append([])
        
        # Preguntas y respuestas
        row_num = 4
        for i, pregunta in enumerate(estadisticas['preguntas'], 1):
            # Encabezado de pregunta
            ws.merge_cells(f'A{row_num}:D{row_num}')
            ws[f'A{row_num}'] = f"{i}. {pregunta['texto_pregunta']} (Tipo: {pregunta['tipo']})"
            ws[f'A{row_num}'].font = Font(bold=True, color="366092")
            ws[f'A{row_num}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row_num += 1
            
            if pregunta['tipo'] == 'opcion_multiple':
                # Tabla para opciones múltiples
                ws[f'A{row_num}'] = "Opción"
                ws[f'B{row_num}'] = "Cantidad"
                ws[f'C{row_num}'] = "Porcentaje"
                
                for cell in ['A', 'B', 'C']:
                    ws[f'{cell}{row_num}'].font = Font(bold=True)
                    ws[f'{cell}{row_num}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                
                row_num += 1
                
                total_respuestas_pregunta = pregunta['total_respuestas']
                for opcion, cantidad in pregunta['conteo_opciones'].items():
                    porcentaje = (cantidad / total_respuestas_pregunta * 100) if total_respuestas_pregunta > 0 else 0
                    
                    ws[f'A{row_num}'] = opcion
                    ws[f'B{row_num}'] = cantidad
                    ws[f'C{row_num}'] = f"{porcentaje:.1f}%"
                    row_num += 1
            else:
                # Respuestas de texto
                ws[f'A{row_num}'] = "Respuestas:"
                ws[f'A{row_num}'].font = Font(bold=True)
                row_num += 1
                
                for respuesta in pregunta.get('respuestas', []):
                    ws[f'A{row_num}'] = respuesta
                    row_num += 1
            
            # Espacio entre preguntas
            row_num += 1
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Guardar en memoria
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def generar_pdf(estadisticas: dict, nombre_archivo: str = "reporte.pdf"):
        """Genera reporte en formato PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = styles['Heading1']
        title_style.alignment = 1  # Centrado
        title = Paragraph(f"Reporte de Encuesta: {estadisticas['titulo_encuesta']}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Información general
        info_style = styles['Normal']
        info_text = f"Total de respuestas: {estadisticas['total_respuestas']}"
        info_paragraph = Paragraph(info_text, info_style)
        elements.append(info_paragraph)
        elements.append(Spacer(1, 0.3*inch))
        
        # Preguntas y respuestas
        for i, pregunta in enumerate(estadisticas['preguntas'], 1):
            # Título de pregunta
            pregunta_style = styles['Heading2']
            pregunta_text = f"{i}. {pregunta['texto_pregunta']} (Tipo: {pregunta['tipo']})"
            pregunta_paragraph = Paragraph(pregunta_text, pregunta_style)
            elements.append(pregunta_paragraph)
            elements.append(Spacer(1, 0.1*inch))
            
            if pregunta['tipo'] == 'opcion_multiple':
                # Tabla para opciones múltiples
                data = [['Opción', 'Cantidad', 'Porcentaje']]
                total_respuestas_pregunta = pregunta['total_respuestas']
                
                for opcion, cantidad in pregunta['conteo_opciones'].items():
                    porcentaje = (cantidad / total_respuestas_pregunta * 100) if total_respuestas_pregunta > 0 else 0
                    data.append([opcion, str(cantidad), f"{porcentaje:.1f}%"])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D9E1F2')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
            else:
                # Respuestas de texto
                respuestas_style = styles['Normal']
                for respuesta in pregunta.get('respuestas', []):
                    respuesta_paragraph = Paragraph(f"• {respuesta}", respuestas_style)
                    elements.append(respuesta_paragraph)
            
            elements.append(Spacer(1, 0.2*inch))
        
        # Generar PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer